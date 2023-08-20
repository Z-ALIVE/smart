from apinfvi import eSight
from DateTime import DateTime
from shutil import copyfile
import datetime
#from assetpivot import *
import vars
import pandas as pd
import json
import os, sys
import glob
import numpy as np
import openpyxl
import win32com.client as win32
import win32api
import win32con
import time
import shutil
import timeit

class ServerDetails():

    def __init__(self):
        self.fields = ["location", "name", "ipAddress", "dn"]
        self.fields2 = ["ipAddress", "bmcMacAddr", "realTimePower", "MemoryCapacity", "cpuCores", "cpuNums", "productSn", "dnsSource"]
        self.jsondata_sd = '//10.150.20.30/SIM Working Directory/Scripts/Asset Inventory/Project Pivot/Script/pivot_asset/asset_pivot_serverdetails.json' #//10.150.20.30/SIM Working Directory/Scripts/Asset Inventory/Project Pivot/Script/pivot_asset/

    def flat_data(self, serverdetails):
        '''
            serverdetails: decoded raw json file output from assetpivot
            :return all server details encoded in json format
        '''
        count_blade = []
        server_details = []
        no_sites = len(serverdetails)
        for i in range(no_sites):
            no_blades = len(serverdetails[i])
            count_blade.append(no_blades)
            for j in range(no_blades):
                out = serverdetails[i][j]
                serverdetails.append(out)
        flat_data = json.dumps(serverdetails[12:], indent=4)
        return flat_data


    def df_ServerDetails(self):
        '''
            :return all server details in DataFrame format
        '''
        json_data = open(self.jsondata_sd, 'r')
        serverdetails = json.load(json_data)
        out = self.flat_data(serverdetails)

        with open('out_sd.json', 'w+') as file:
            file.write(out)

        df_ServerDetails = pd.read_json('//10.150.20.30/SIM Working Directory/Scripts/Asset Inventory/Project Pivot/Script/pivot_report/out_sd.json')

        return df_ServerDetails


    def df_blade(self):
        '''
            :return: all blade details
        '''
        ServerDetails = self.df_ServerDetails()
        df_blade = pd.json_normalize(ServerDetails[0], "board", self.fields2, record_prefix='_', errors='ignore')
        return df_blade

    def df_cpu(self):
        '''
            :return: all Server CPU details in DataFrame
        '''
        ServerDetails = self.df_ServerDetails()
        df_cpu = pd.json_normalize(ServerDetails[0], "CPU", self.fields, record_prefix='_')
        return df_cpu

    def df_disk(self):
        '''
            :return: all Server Disk details in DataFrame
        '''
        ServerDetails = self.df_ServerDetails()
        df_disk = pd.json_normalize(ServerDetails[0], "Disk", self.fields, record_prefix='_')
        return df_disk

    def df_memory(self):
        '''
            :return: all Server Memory details in DataFrame
        '''
        ServerDetails = self.df_ServerDetails()
        df_memory = pd.json_normalize(ServerDetails[0], "Memory", self.fields, record_prefix='_')
        return df_memory

    def df_mezz(self):
        '''
            :return: all Server Mezz details in DataFrame
        '''
        ServerDetails = self.df_ServerDetails()
        df_mezz = pd.json_normalize(ServerDetails[0], "Mezz", self.fields, record_prefix='_')
        return df_mezz

    def df_netcard(self):
        '''
            :return: all Server NetCard details in DataFrame
        '''
        ServerDetails = self.df_ServerDetails()
        df_netcard = pd.json_normalize(ServerDetails[0], "NetCard", self.fields, record_prefix='_')
        return df_netcard

    def df_networkcard(self):
        '''
            :return: all Server NetworkCard details in DataFrame
        '''
        ServerDetails = self.df_ServerDetails()
        df_networkcard = pd.json_normalize(ServerDetails[0], "NetworkCard", self.fields, record_prefix='_')
        return df_networkcard

    def df_raid(self):
        '''
            :return: all Server RAID details in DataFrame
        '''
        ServerDetails = self.df_ServerDetails()
        df_raid = pd.json_normalize(ServerDetails[0], "RAID", self.fields, record_prefix='_')
        return df_raid

class ServerList():

    def __init__(self):
        self.fields = ["location", "serverName", "ipAddress", "dn", "status", "version",  "serverModel", "productSN", "manufacturer", "uuid"]
        self.jsondata_sl = '//10.150.20.30/SIM Working Directory/Scripts/Asset Inventory/Project Pivot/Script/pivot_asset/asset_pivot_serverlist.json'

    def flat_data(self, serverlist):
        '''
            serverdetails: decoded raw json file output from assetpivot
            :return all server details encoded in json format
        '''
        count_e9000 = []
        server_list = []
        no_sites = len(serverlist)
        for i in range(no_sites):
            no_e9000 = len(serverlist[i])
            count_e9000.append(no_e9000)
            for j in range(no_e9000):
                out = serverlist[i][j]
                server_list.append(out)
        flat_data = json.dumps(server_list, indent=4)
        return flat_data

    def df_ServerList(self):
        '''
            :return: all server list in json format
        '''
        json_data = open(self.jsondata_sl, 'r')
        serverlist = json.load(json_data)
        out_json = self.flat_data(serverlist)
        out = json.loads(out_json)
        return out

    def df_childBlades(self):
        '''
            :return: all blade server list in DataFrame
        '''
        ServerList = self.df_ServerList()
        df_childBlades = pd.json_normalize(ServerList, "childBlades", self.fields, errors='ignore', record_prefix='_')
        return df_childBlades

    def df_switchBoard(self):
        '''
            :return: all switchboard list in DataFrame
        '''
        ServerList = self.df_ServerList()
        df_switchBoard = pd.json_normalize(ServerList, "switchBoard", self.fields, errors='ignore', record_prefix='_')
        return df_switchBoard

class StorageDetails():

    def __init__(self):
        self.jsondata_sd = '//10.150.20.30/SIM Working Directory/Scripts/Asset Inventory/Project Pivot/Script/pivot_asset/asset_pivot_storagedisklist.json'

    def flat_data(self,storagedetails):
        '''

        :param storagedisklist:
        :return:
        '''
        count_storage = []
        storage_list = []
        no_sites = len(storagedetails)
        for i in range(no_sites):
            no_storage = len(storagedetails[i])
            count_storage.append(no_storage)
            for j in range(no_storage):
                out = storagedetails[i][j]
                storage_list.append(out)
            flat_data = json.dumps(storage_list, indent=4)
        return flat_data

    def df_StorageDetails(self):
        json_data = open(self.jsondata_sd,'r')
        storagedetails = json.load(json_data)
        out_json = self.flat_data(storagedetails)
        out = json.loads(out_json)
        return out

    def df_disk(self):
        StorageDetails = self.df_StorageDetails()
        df_disk = pd.json_normalize(StorageDetails,  errors='ignore', record_prefix='_')
        return df_disk

class StorageList():

    def __init__(self):
        self.jsondata_sl = '//10.150.20.30/SIM Working Directory/Scripts/Asset Inventory/Project Pivot/Script/pivot_asset/asset_pivot_storagelist.json'

    def flat_data(self,storagelist):
        '''
        :param storagedisklist:
        :return:
        '''
        count_storage = []
        storage_list = []
        no_sites = len(storagelist)
        for i in range(no_sites):
            no_storage = len(storagelist[i])
            count_storage.append(no_storage)
            for j in range(no_storage):
                out = storagelist[i][j]
                storage_list.append(out)
            flat_data = json.dumps(storage_list, indent=4)
        return flat_data

    def df_StorageList(self):
        json_data = open(self.jsondata_sl,'r')
        storagelist = json.load(json_data)
        out_json = self.flat_data(storagelist)
        out = json.loads(out_json)
        return out

    def df_storage(self):
        StorageList = self.df_StorageList()
        df_storage = pd.json_normalize(StorageList,  errors='ignore', record_prefix='_')
        return df_storage

class NetworkList():

    def __init__(self):
        self.jsondata_nl = '//10.150.20.30/SIM Working Directory/Scripts/Asset Inventory/Project Pivot/Script/pivot_asset/asset_pivot_networklist.json'

    def flat_data(self,networklist):
        '''
        :param networklist:
        :return:
        '''
        count_network = []
        network_list = []
        no_sites = len(networklist)
        for i in range(no_sites):
            no_network = len(networklist[i])
            count_network.append(no_network)
            for j in range(no_network):
                out = networklist[i][j]
                network_list.append(out)
            flat_data = json.dumps(network_list, indent=4)
        return flat_data

    def df_NetworkList(self):
        json_data = open(self.jsondata_nl,'r')
        networklist = json.load(json_data)
        out_json = self.flat_data(networklist)
        out = json.loads(out_json)
        return out

    def df_network(self):
        NetworkList = self.df_NetworkList()
        df_networklist = pd.json_normalize(NetworkList,  errors='ignore', record_prefix='_')
        return df_networklist

class PivotReport():
    '''
    Refactoring DataFrames
    '''
    serverdetails = ServerDetails() # Calling instance of Class ServerDetails
    serverlist = ServerList()  # Calling instance of Class ServerList
    storagedetails = StorageDetails() # Calling instance of Class StorageDetails
    storagelist = StorageList() # Calling instance of Class StorageList
    networklist = NetworkList() # Calling instance of Class NetworkList

    def __init__(self):
        self.array = 'Location', 'Server IP Address', 'Server Name', 'Blade No', 'Blade Name', 'Blade IP Address'
        self.lookupval = 'Blade IP Address'

    ##########################################################
    '''~~~~~~~~~~~~~~~~~~~Blade Details~~~~~~~~~~~~~~~~~~~~'''
    ##########################################################

    def childBlades(self):
        '''
        Server Details from ServerList()
        '''
        childBlades = self.serverlist.df_childBlades()
        childBlades = childBlades.assign()

        childBlades = childBlades.assign(HealthState=childBlades['_state'].map(lambda x: {'0': 'Normal', '1': 'Offline', '2': 'Unknown', 'Others': 'Faulty'}.get(x, '--')))
        childBlades = childBlades.assign(ServerState=childBlades['status'].map(lambda x: {'0': 'Normal', '1': 'Offline', '2': 'Unknown', 'Others': 'Faulty'}.get(x, '--')))
        childBlades = childBlades.assign(CPUHealthState=childBlades['_cpuHealthState'].map(lambda x: {'0': 'Normal', '1': 'Absent', '2': 'Unknown', 'Others': 'Faulty'}.get(x, '--')))
        childBlades = childBlades.assign(MemoryHealthState=childBlades['_memoryHealthState'].map(lambda x: {'0': 'Normal', '1': 'Absent', '2': 'Unknown', 'Others': 'Faulty'}.get(x, '--')))

        childBlades = childBlades[['ipAddress', 'serverName', '_location', '_name', '_ipAddress', '_type', 'serverModel', '_dn', 'HealthState', '_version', 'dn', 'ServerState', 'version', 'productSN', 'CPUHealthState', 'MemoryHealthState', 'manufacturer', 'uuid']]
        childBlades.insert(0, 'Location', None)
        childBlades = childBlades.set_axis(['Location', 'Server IP Address', 'Server Name', 'Blade No', 'Blade Name', 'Blade IP Address', 'Type', 'Server Model', 'Blade DN', 'Health State', 'Blade Version', 'Server DN', 'Server State', 'Server Version', 'Sever Product SN', 'CPU Health State', 'Memory Health State', 'Manufacturer', 'UUID'], axis='columns')

        childBlades_count = len(childBlades['Server Name'])
        for i in range(childBlades_count):
            if childBlades.loc[i]['Server Name'][:3] == 'SFU':
                childBlades.loc[i]['Location'] = 'SFLU'
            elif childBlades.loc[i]['Server Name'][:3] == 'ANG':
                childBlades.loc[i]['Location'] = 'Clark'
            elif childBlades.loc[i]['Server Name'][:3] == 'LCN':
                childBlades.loc[i]['Location'] = 'Lucena'
            elif childBlades.loc[i]['Server Name'][:3] == 'BAT':
                childBlades.loc[i]['Location'] = 'Batangas'
            elif childBlades.loc[i]['Server Name'][:3] == 'GHL':
                childBlades.loc[i]['Location'] = 'Greenhills'
            elif childBlades.loc[i]['Server Name'][:3] == 'SPC':
                childBlades.loc[i]['Location'] = 'Sampaloc'
            elif childBlades.loc[i]['Server Name'][:3] == 'MKT':
                childBlades.loc[i]['Location'] = 'Makati'
            elif childBlades.loc[i]['Server Name'][:3] == 'PQE':
                childBlades.loc[i]['Location'] = 'Parañaque'
            elif childBlades.loc[i]['Server Name'][:3] == 'CEB':
                childBlades.loc[i]['Location'] = 'Cebu'
            elif childBlades.loc[i]['Server Name'][:3] == 'ILO':
                childBlades.loc[i]['Location'] = 'Iloilo'
            elif childBlades.loc[i]['Server Name'][:3] == 'CDO':
                childBlades.loc[i]['Location'] = 'CDO'
            elif childBlades.loc[i]['Server Name'][:3] == 'DAV':
                childBlades.loc[i]['Location'] = 'Davao'

        return childBlades

    ################################################################
    '''~~~~~~~~~~~~~~~~~~~Switchboard Details~~~~~~~~~~~~~~~~~~~~'''
    ################################################################

    def switchBoard(self):
        '''
        CX Board Details from ServerList()
        '''
        switchBoard = self.serverlist.df_switchBoard()
        switchBoard = switchBoard.assign(HealthState=switchBoard['_healthState'].map(lambda x: {'0': 'Normal', '1': 'Offline', '2': 'Unknown', 'Others': 'Faulty'}.get(x, '--')))
        switchBoard = switchBoard.assign(PresentState=switchBoard['_presentState'].map(lambda x: {'0': 'Absent', '1': 'Present', '2': 'Unknown', 'Others': 'Faulty'}.get(x, '--')))
        switchBoard = switchBoard.assign(Type=switchBoard['_type'].map(lambda x: {0: 'Mainboard', 1: 'Switch Module'}.get(x, '--')))
        switchBoard = switchBoard[['location', 'ipAddress', 'serverName', '_name', '_ipAddress', 'HealthState', 'PresentState', '_dn',  'Type', '_manufacture', '_sn', '_partNumber', '_moId', '_manuTime', '_uuid']]
        switchBoard = switchBoard.set_axis(['Location', 'Server IP Address', 'Server Name', 'CX Board Name', 'CX Board IP Address', 'Health State', 'Present State', 'DN',  'Type', 'Manufacture', 'SN', 'Part Number', 'MOID', 'ManuTime', 'UUID'], axis='columns')

        switchBoard_count = len(switchBoard['Server Name'])
        for i in range(switchBoard_count):
            if switchBoard.loc[i]['Server Name'][:3] == 'SFU':
                switchBoard.loc[i]['Location'] = 'SFLU'
            elif switchBoard.loc[i]['Server Name'][:3] == 'ANG':
                switchBoard.loc[i]['Location'] = 'Clark'
            elif switchBoard.loc[i]['Server Name'][:3] == 'LCN':
                switchBoard.loc[i]['Location'] = 'Lucena'
            elif switchBoard.loc[i]['Server Name'][:3] == 'BAT':
                switchBoard.loc[i]['Location'] = 'Batangas'
            elif switchBoard.loc[i]['Server Name'][:3] == 'GHL':
                switchBoard.loc[i]['Location'] = 'Greenhills'
            elif switchBoard.loc[i]['Server Name'][:3] == 'SPC':
                switchBoard.loc[i]['Location'] = 'Sampaloc'
            elif switchBoard.loc[i]['Server Name'][:3] == 'MKT':
                switchBoard.loc[i]['Location'] = 'Makati'
            elif switchBoard.loc[i]['Server Name'][:3] == 'PQE':
                switchBoard.loc[i]['Location'] = 'Parañaque'
            elif switchBoard.loc[i]['Server Name'][:3] == 'CEB':
                switchBoard.loc[i]['Location'] = 'Cebu'
            elif switchBoard.loc[i]['Server Name'][:3] == 'ILO':
                switchBoard.loc[i]['Location'] = 'Iloilo'
            elif switchBoard.loc[i]['Server Name'][:3] == 'CDO':
                switchBoard.loc[i]['Location'] = 'CDO'
            elif switchBoard.loc[i]['Server Name'][:3] == 'DAV':
                switchBoard.loc[i]['Location'] = 'Davao'

        return switchBoard


    ##########################################################
    '''~~~~~~~~~~~~~~~~~~~Blade Details~~~~~~~~~~~~~~~~~~~~'''
    ##########################################################

    def server(self):
        '''
            :return: Server in Data Frame
        '''
        blade = self.serverdetails.df_blade() # details from server details
        blade = blade[['ipAddress', 'bmcMacAddr', 'realTimePower', 'MemoryCapacity', 'cpuCores', 'cpuNums', 'productSn', 'dnsSource']]
        blade = blade.set_axis(['Blade IP Address', 'Blade MAC Address', 'Real Time Power', 'Memory Capacity', 'CPU Cores', 'CPU Nums', 'Blade Product SN', 'DNS Source'], axis='columns')
        childblades = self.childBlades()
        blade = pd.merge(childblades, blade, on=self.lookupval, how='left')
        sheet_server = blade[['Location', 'Server IP Address', 'Server Name', 'Blade No',
               'Blade Name', 'Blade IP Address', 'Type', 'Server Model', 'Blade DN',
               'Health State', 'Blade Version', 'Server DN', 'Server State',
               'Server Version', 'Sever Product SN', 'CPU Health State',
               'Memory Health State', 'Blade MAC Address',
               'Real Time Power', 'Memory Capacity', 'CPU Cores', 'CPU Nums',
               'Blade Product SN', 'DNS Source', 'Manufacturer', 'UUID']]
        return sheet_server

    def array_lookup(self):
        '''
            :return: an array
        '''
        server = self.server()
        array_lookup = server[['Location', 'Server IP Address', 'Server Name', 'Blade No', 'Blade Name', 'Blade IP Address']]
        return array_lookup

    def cpu(self):
        '''
            :return: CPU Details in Data Frame
        '''
        cpu = self.serverdetails.df_cpu()
        cpu = cpu.assign(HealthState=cpu['_healthState'].map(lambda x: {'0': 'Normal', '1': 'Offline', '2': 'Unknown', 'Others': 'Faulty'}.get(x, '--')))
        cpu = cpu.assign(PresentState=cpu['_presentState'].map(lambda x: {'0': 'Absent', '1': 'Present', '2': 'Unknown', 'Others': 'Faulty'}.get(x, '--')))
        cpu = cpu.assign(EnableState=cpu['_enableState'].map(lambda  x: {'1':'Enabled', '2':'Disabled'}.get(x, '--')))
        cpu = cpu[['ipAddress', 'dn', '_name', 'HealthState', 'PresentState', 'EnableState', '_manufacture', '_model', '_frequency', '_processorId','_totalCores', '_totalThreads', '_l1CacheKiB', '_l2CacheKiB', '_l3CacheKiB', '_partNumber', '_uuid', '_moId']]
        cpu = cpu.set_axis(['Blade IP Address', 'Blade DN', 'CPU Name', 'Health State', 'Present State', 'Enable State', 'Vendor', 'CPU Model',
             'CPU Frequency', 'CPU Processor ID', 'CPU Total Cores', 'CPU Total Threads', 'CPU L1 Cache', 'CPU L2 Cache',
             'CPU L3 Cache', 'CPU BOM Code', 'UUID', 'MOID'], axis='columns')
        array = self.array_lookup()
        cpu = pd.merge(cpu, array, on=self.lookupval, how='left')
        sheet_cpu = cpu[
            ['Location', 'Server IP Address', 'Server Name', 'Blade No', 'Blade Name', 'Blade IP Address',
             'Blade DN', 'CPU Name', 'Health State', 'Present State', 'Enable State', 'Vendor', 'CPU Model',
             'CPU Frequency', 'CPU Processor ID', 'CPU Total Cores', 'CPU Total Threads', 'CPU L1 Cache', 'CPU L2 Cache',
             'CPU L3 Cache', 'CPU BOM Code', 'UUID', 'MOID']]
        return sheet_cpu

    def disk(self):
        '''
            :return: DISK Details in Data Frame
        '''
        disk = self.serverdetails.df_disk()
        disk = disk.assign(HealthState=disk['_healthState'].map(lambda x: {'0': 'Normal', '1': 'Offline', '2': 'Unknown', 'Others': 'Faulty'}.get(x, 'NA')))
        disk = disk.assign(PresentState=disk['_presentState'].map(lambda x: {'0': 'Absent', '1': 'Present', '2': 'Unknown', 'Others': 'Faulty'}.get(x, 'NA')))
        disk = disk[['ipAddress', 'dn', '_name', '_capacity', 'HealthState', 'PresentState', '_model', '_manufacturer', '_moId', '_uuid',]]
        disk = disk.set_axis(['Blade IP Address', 'Blade DN', 'Disk Name', 'Disk Capacity', 'Health State', 'Present State', 'Disk Model', 'Manufacturer', 'MOID', 'UUID'],axis='columns')
        array = self.array_lookup()
        disk = pd.merge(disk, array, on=self.lookupval, how='left')
        sheet_disk = disk[
            ['Location', 'Server IP Address', 'Server Name', 'Blade No', 'Blade Name', 'Blade IP Address',
             'Blade DN', 'Disk Name', 'Disk Capacity', 'Health State',
             'Present State', 'Disk Model', 'Manufacturer', 'MOID', 'UUID']]
        return sheet_disk

    def memory(self):
        '''
            :return: MEMORY Details in Data Frame
        '''
        memory = self.serverdetails.df_memory()
        memory = memory.assign(HealthState=memory['_healthState'].map(lambda x: {'0':'Normal', '1':'Offline', '2':'Unknown', 'Others':'Faulty'}.get(x, 'NA')))
        memory = memory.assign(PresentState=memory['_presentState'].map(lambda x: {'0':'Absent', '1':'Present', '2':'Unknown', 'Others':'Faulty'}.get(x, 'NA')))
        memory = memory[['ipAddress', 'dn', '_name', '_capacity', '_frequency', '_dataWidthBits',  '_rankCount',  'HealthState', 'PresentState','_manufacture', '_memoryType', '_minVoltage', '_partNumber', '_serialNumber', '_technology', '_moId', '_uuid']]
        memory = memory.set_axis(
            ['Blade IP Address', 'Blade DN', 'Memory Name', 'Memory Capacity', 'Frequency', 'Bit Width', 'Rank Count',  'Health State',
             'Present State', 'Manufacturer', 'Mmeory Type', 'Minimum Voltage', 'Part Number', 'Product SN', 'Technology', 'MOID', 'UUID'], axis='columns')
        array = self.array_lookup()
        memory = pd.merge(memory,array, on=self.lookupval, how='left')
        sheet_memory = memory[
            ['Location', 'Server IP Address', 'Server Name', 'Blade No', 'Blade Name', 'Blade IP Address',
             'Blade DN', 'Memory Name', 'Memory Capacity', 'Frequency', 'Bit Width', 'Rank Count',  'Health State',
             'Present State', 'Manufacturer', 'Mmeory Type', 'Minimum Voltage', 'Part Number', 'Product SN', 'Technology', 'MOID', 'UUID']]
        return sheet_memory

    def mezz(self):
        '''
            :return: MEZZ Details in Data Frame
        '''
        mezz = self.serverdetails.df_mezz()
        mezz = mezz.assign(HealthState=mezz['_mezzHealthStatus'].map(lambda x: {'0': 'Normal', '1': 'Offline', '2': 'Unknown', 'Others': 'Faulty'}.get(x, 'NA')))
        mezz = mezz.assign(PresentState=mezz['_presentState'].map(lambda x: {'0': 'Absent', '1': 'Present', '2': 'Unknown', 'Others': 'Faulty'}.get(x, 'NA')))
        mezz = mezz[['ipAddress','dn', '_name', '_mezzInfo', '_mezzLocation', '_mezzMac', 'HealthState','PresentState', '_moId', '_uuid']]
        mezz = mezz.set_axis(['Blade IP Address', 'Blade DN', 'Mezz Name', 'Mezz Info', 'Mezz Location', 'Mezz MAC', 'Health State', 'Present State', 'MOID', 'UUID'], axis='columns')
        array = self.array_lookup()
        mezz = pd.merge(mezz, array, on=self.lookupval, how='left')
        sheet_mezz = mezz[
            ['Location', 'Server IP Address', 'Server Name', 'Blade No', 'Blade Name', 'Blade IP Address',
             'Blade DN', 'Mezz Name', 'Mezz Info', 'Mezz Location', 'Mezz MAC', 'Health State', 'Present State', 'MOID', 'UUID']]
        return sheet_mezz

    def netcard(self):
        '''
            :return: NETCARD Details in Data Frame
        '''
        netcard = self.serverdetails.df_netcard()
        netcard = netcard.assign(HealthState=netcard['_healthState'].map(lambda x: {'0': 'Normal', '1': 'Offline', '2': 'Unknown', 'Others': 'Faulty'}.get(x, 'NA')))
        netcard = netcard[['ipAddress', 'dn', '_deviceName',  '_cardName', 'HealthState',  '_chipModel', '_chipManufacturer',  '_cardModel', '_cardManufacturer', '_driverName', '_driverVersion', '_firmwarePkgVersion', '_moId', '_uuid']]
        netcard = netcard.set_axis(
            ['Blade IP Address', 'Blade DN', 'Device Name', 'Card Name', 'Health State', 'Chip Model', 'Chip Manufacturer', 'Card Model',
             'Card Manufacturer', 'Driver Name', 'Driver Version', 'Firmware Version', 'MOID', 'UUID'], axis='columns')
        array = self.array_lookup()
        netcard = pd.merge(netcard, array, on=self.lookupval, how='left')
        sheet_netcard = netcard[
            ['Location', 'Server IP Address', 'Server Name', 'Blade No', 'Blade Name', 'Blade IP Address',
             'Blade DN', 'Device Name', 'Card Name', 'Health State', 'Chip Model', 'Chip Manufacturer', 'Card Model',
             'Card Manufacturer', 'Driver Name', 'Driver Version', 'Firmware Version', 'MOID', 'UUID']]
        return sheet_netcard

    def netport(self):
        '''
            :return: NETPORT Details in Data Frame
        '''
        netport = self.serverdetails.df_networkcard()
        netport = netport[['ipAddress', 'dn',  '_netWorkCardName', '_portName',  '_macAdress', '_ipv4Addresses', '_ipv6Addresses', '_ipv6DefaultGateway', '_linkStatus', '_moId', '_uuid']]
        netport = netport.set_axis(
            ['Blade IP Address', 'Blade DN', 'Network Card Name', 'Port Name', 'MAC Address', 'IPv4 Address', 'IPv6 Address',
             'IPv6 Default Gateway', 'Link Status', 'MOID', 'UUID'], axis='columns')
        array = self.array_lookup()
        netport = pd.merge(netport, array, on=self.lookupval, how='left')
        sheet_netport = netport[
            [ 'Location', 'Server IP Address', 'Server Name', 'Blade No', 'Blade Name', 'Blade IP Address',
              'Blade DN', 'Network Card Name', 'Port Name', 'MAC Address', 'IPv4 Address', 'IPv6 Address',
              'IPv6 Default Gateway', 'Link Status', 'MOID', 'UUID']]
        return sheet_netport

    def raid(self):
        '''
            :return: RAID Details in Data Frame
        '''
        raid = self.serverdetails.df_raid()
        raid = raid.assign(HealthState=raid['_healthState'].map(lambda x: {'0': 'Normal', '-1': 'Offline', '-2': 'Unknown', 'Others': 'Faulty'}.get(x, '--')))
        raid = raid.assign(InterfaceType=raid['_interfaceType'].map(lambda x: {'1': 'SPI', '2': 'SAS_3G', '3': 'SATA_1.5G', '4': 'SATA_3G','5': 'SAS_6G', '6':'SAS_12G', '255':'Unknown'}.get(x, '--')))
        raid = raid[['ipAddress', 'dn', '_name', 'HealthState', '_bbuType',  'InterfaceType', '_raidLevel', '_raidType', '_moId', '_uuid']]
        raid = raid.set_axis(['Blade IP Address', 'Blade DN', 'RAID Name', 'Health State', 'BBU Type', 'Interface Type', 'RAID Level', 'RAID Type', 'MOID', 'UUID'], axis='columns')
        array = self.array_lookup()
        raid = pd.merge(raid, array, on=self.lookupval, how='left')
        sheet_raid = raid[
            [ 'Location', 'Server IP Address', 'Server Name', 'Blade No', 'Blade Name', 'Blade IP Address',
              'Blade DN', 'RAID Name', 'Health State', 'BBU Type', 'Interface Type', 'RAID Level', 'RAID Type', 'MOID', 'UUID']]
        return sheet_raid

    def storage(self):
        storl = self.storagelist.df_storage()
        storl = storl.assign(HealthState=storl['status'].map(lambda x: {'0': 'Normal', '-1': 'Offline', '1': 'Degraded','2': 'Faulty', '4': 'Unknown', '9':'Unmanaged'}.get(x, '--')))
        storl = storl[['location', 'name', 'ipAddress', 'dn','HealthState', 'recentUpdateTime', 'totalCap',  'usablecapacity', 'usedCap', 'mapedluncapacity', 'unmapedluncapacity', 'hotdiskcapacity', 'hotfreecapacity', 'hotusedcapacity', 'version', 'productSN', 'manufacturer']]
        storl = storl.set_axis(['Location', 'Storage Name', 'Storage IP Address','DN', 'Health State', 'Recent Update Time', 'Total Capacity', 'Usable Capacity', 'Used Capacity', 'Maped LUN Capacity', 'Unmaped LUN Capacity', 'Hot Disk Capacity', 'Hot Free Capacity', 'Hot Used Capacity', 'Version', 'Product SN', 'Manufacturer'], axis='columns')

        storl_count = len(storl['Storage Name'])
        for i in range(storl_count):
            if storl.loc[i]['Storage Name'][:3] == 'SFU':
                storl.loc[i]['Location'] = 'SFLU'
            elif storl.loc[i]['Storage Name'][:3] == 'ANG':
                storl.loc[i]['Location'] = 'Clark'
            elif storl.loc[i]['Storage Name'][:3] == 'LCN':
                storl.loc[i]['Location'] = 'Lucena'
            elif storl.loc[i]['Storage Name'][:3] == 'BAT':
                storl.loc[i]['Location'] = 'Batangas'
            elif storl.loc[i]['Storage Name'][:3] == 'GHL':
                storl.loc[i]['Location'] = 'Greenhills'
            elif storl.loc[i]['Storage Name'][:3] == 'SPC':
                storl.loc[i]['Location'] = 'Sampaloc'
            elif storl.loc[i]['Storage Name'][:3] == 'MKT':
                storl.loc[i]['Location'] = 'Makati'
            elif storl.loc[i]['Storage Name'][:3] == 'PQE':
                storl.loc[i]['Location'] = 'Parañaque'
            elif storl.loc[i]['Storage Name'][:3] == 'CEB':
                storl.loc[i]['Location'] = 'Cebu'
            elif storl.loc[i]['Storage Name'][:3] == 'ILO':
                storl.loc[i]['Location'] = 'Iloilo'
            elif storl.loc[i]['Storage Name'][:3] == 'CDO':
                storl.loc[i]['Location'] = 'CDO'
            elif storl.loc[i]['Storage Name'][:3] == 'DAV':
                storl.loc[i]['Location'] = 'Davao'

        return storl

    def storagedisk(self):
        storl = self.storage()
        storl = storl[['Location', 'Storage Name', 'Storage IP Address','DN']]

        stord = self.storagedetails.df_disk()
        stord = stord.assign(HealthStatus=stord['healthStatus'].map(lambda x: {0: 'Normal', 3: 'to be faulty', 2: 'Faulty', 'Others': 'Unkown'}.get(x, '--')))
        stord = stord.assign(RunningStatus=stord['runningStatus'].map(lambda x: {0: 'Normal', -1: 'Offline', 2: 'Faulty', 'Others': 'Unkown'}.get(x, '--')))
        stord = stord[['dn', 'diskPos', 'totalCapacity', 'HealthStatus', 'logicalType', 'physicalModel', 'poolId', 'RunningStatus', 'moId', ]]
        stord = stord.set_axis(['DN', 'Disk Position', 'Total Capacity', 'Health Status', 'Logical Type', 'Physical Model', 'Pool ID','Running Status', 'MOID'], axis='columns')

        stord = pd.merge(storl, stord, on='DN', how='left')

        return stord

    def network(self):
        net = self.networklist.df_network()
        net = net.assign(NEState=net['nestate'].map(lambda x: {0: 'Not Detected', 1: 'Online', 2: 'Offline', 3: 'Unkown'}.get(x, '--')))
        net = net[['neposition', 'nename', 'neip', 'netype', 'NEState', 'nemac', 'neosversion', 'nepatchversion', 'necategory', 'nedn', 'negroupname', 'nevendorname', 'timezoneid',  'version', 'nedescribe']]
        net = net.set_axis(['Location', 'Device Name', 'Device IP Address', 'Device Type', 'Device State', 'Device MAC', 'OS Version', 'Patch Version', 'Device Category', 'DN', 'Group Name', 'Vendor', 'Time Zone', 'Version', 'Description'], axis='columns')

        net_count = len(net['Device Name'])
        for i in range(net_count):
            if net.loc[i]['Device Name'][:3] == 'SFU':
                net.loc[i]['Location'] = 'SFLU'
            elif net.loc[i]['Device Name'][:3] == 'ANG':
                net.loc[i]['Location'] = 'Clark'
            elif net.loc[i]['Device Name'][:3] == 'LCN':
                net.loc[i]['Location'] = 'Lucena'
            elif net.loc[i]['Device Name'][:3] == 'BAT':
                net.loc[i]['Location'] = 'Batangas'
            elif net.loc[i]['Device Name'][:3] == 'GHL':
                net.loc[i]['Location'] = 'Greenhills'
            elif net.loc[i]['Device Name'][:3] == 'SPC':
                net.loc[i]['Location'] = 'Sampaloc'
            elif net.loc[i]['Device Name'][:3] == 'MKT':
                net.loc[i]['Location'] = 'Makati'
            elif net.loc[i]['Device Name'][:3] == 'PQE':
                net.loc[i]['Location'] = 'Parañaque'
            elif net.loc[i]['Device Name'][:3] == 'CEB':
                net.loc[i]['Location'] = 'Cebu'
            elif net.loc[i]['Device Name'][:3] == 'ILO':
                net.loc[i]['Location'] = 'Iloilo'
            elif net.loc[i]['Device Name'][:3] == 'CDO':
                net.loc[i]['Location'] = 'CDO'
            elif net.loc[i]['Device Name'][:3] == 'DAV':
                net.loc[i]['Location'] = 'Davao'
            elif net.loc[i]['Device IP Address'] == '10.90.206.84':
                net.loc[i]['Location'] = 'SFLU'
            elif net.loc[i]['Device IP Address'] == '10.90.113.84':
                net.loc[i]['Location'] = 'Clark'
            elif net.loc[i]['Device IP Address'] == '10.90.222.84':
                net.loc[i]['Location'] = 'Lucena'
            elif net.loc[i]['Device IP Address'] == '10.90.154.84':
                net.loc[i]['Location'] = 'Batangas'
            elif net.loc[i]['Device IP Address'] == '10.90.130.84':
                net.loc[i]['Location'] = 'Greenhills'
            elif net.loc[i]['Device IP Address'] == '10.90.152.84':
                net.loc[i]['Location'] = 'Sampaloc'
            elif net.loc[i]['Device IP Address'] == '10.90.40.84':
                net.loc[i]['Location'] = 'Makati'
            elif net.loc[i]['Device IP Address'] == '10.90.190.84':
                net.loc[i]['Location'] = 'Parañaque'
            elif net.loc[i]['Device IP Address'] == '10.90.69.84':
                net.loc[i]['Location'] = 'Cebu'
            elif net.loc[i]['Device IP Address'] == '10.90.253.84':
                net.loc[i]['Location'] = 'Iloilo'
            elif net.loc[i]['Device IP Address'] == '10.90.86.84':
                net.loc[i]['Location'] = 'Davao'

        return net

    def pivot_table(self):
        summary = self.summary()

        pivot_table = summary.pivot_table(index=['Location', 'Server Name'], values=['Server CPU', 'Server Disk', 'Server Memory', 'Server MEZZ', 'Server Netport', 'Server Netcard',
                                    'Server RAID'], aggfunc='sum')

        return pivot_table

    def summary_server(self):
        server = self.server()
        cpu = self.cpu()
        disk = self.disk()
        memory = self.memory()
        mezz = self.mezz()
        netport = self.netport()
        netcard = self.netcard()
        raid = self.raid()
        swboard = self.switchBoard()

        '''
            :return summary of Server Components in Data Frame
        '''

        summary = server[['Location', 'Server IP Address', 'Server Name', 'Blade No', 'Blade Name', 'Blade IP Address']]
        #summary = summary.reindex(columns=summary.columns.tolist() + ['Server CPU', 'Server Disk', 'Server Memory', 'Server Netport', 'Server Netcard', 'Server Mezz', 'Server RAID'])

        cpu = cpu.groupby(['Blade IP Address'])[['CPU Name']].count()
        disk = disk.groupby(['Blade IP Address'])[['Disk Name']].count()
        memory = memory.groupby(['Blade IP Address'])[['Memory Name']].count()
        mezz = mezz.groupby(['Blade IP Address'])[['Mezz Name']].count()
        netport = netport.groupby(['Blade IP Address'])[['Network Card Name']].count()
        netcard = netcard.groupby(['Blade IP Address'])[['Card Name']].count()
        raid = raid.groupby(['Blade IP Address'])[['RAID Name']].count()

        summary = pd.merge(summary, cpu, on=self.lookupval, how='left')
        summary = pd.merge(summary, disk, on=self.lookupval, how='left')
        summary = pd.merge(summary, memory, on=self.lookupval, how='left')
        summary = pd.merge(summary, mezz, on=self.lookupval, how='left')
        summary = pd.merge(summary, netport, on=self.lookupval, how='left')
        summary = pd.merge(summary, netcard, on=self.lookupval, how='left')
        summary = pd.merge(summary, raid, on=self.lookupval, how='left')

        summary = summary.set_axis(['Location', 'Server IP Address', 'Server Name', 'Blade No', 'Blade Name', 'Blade IP Address',
                                    'Server CPU', 'Server Disk', 'Server Memory', 'Server MEZZ', 'Server Netport', 'Server Netcard',
                                    'Server RAID'], axis='columns')
        return summary

    def summary_swboard(self):
        swboard = self.switchBoard()
        summary = swboard.groupby(['Location', 'Server IP Address', 'Server Name'])[['CX Board Name']].count()
        return summary

    def summary_storage(self):
        storage = self.storagedisk()
        summary = storage.groupby(['Location','Storage Name', 'Storage IP Address','DN'])[['Disk Position']].count()
        return summary

    def summary_network(self):
        network = self.network()
        summary = network.groupby(['Location'])[['Device Name']].count()
        return summary

    def null(self):
        df = pd.DataFrame(None)
        return df

print("Start Running Script for PIVOT ASSET INVENTORY\n-----------------------------------------------------\n")
start = timeit.default_timer()
current_time = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
print('Start Time: {}'.format (current_time))

writer = pd.ExcelWriter('PIVOT ASSET INVENTORY.xlsx')
pivot_report = PivotReport()

pivot_report.null().to_excel(writer, sheet_name='COMPUTE')
pivot_report.null().to_excel(writer, sheet_name='LEAFSW')
pivot_report.null().to_excel(writer, sheet_name='STORAGE')
pivot_report.null().to_excel(writer, sheet_name='EOR_TOR_LANSW_AC')
pivot_report.summary_server().to_excel(writer, sheet_name='_COMPUTE', merge_cells=False)
pivot_report.summary_swboard().to_excel(writer, sheet_name='_LEAFSW', merge_cells=False)
pivot_report.summary_storage().to_excel(writer, sheet_name='_STORAGE', merge_cells=False)
pivot_report.summary_network().to_excel(writer, sheet_name='_EOR_TOR_LANSW_AC', merge_cells=False)
pivot_report.server().to_excel(writer, sheet_name='Server')
pivot_report.cpu().to_excel(writer, sheet_name='Server CPU')
pivot_report.disk().to_excel(writer, sheet_name='Server Disk')
pivot_report.memory().to_excel(writer, sheet_name='Server Memory')
pivot_report.mezz().to_excel(writer, sheet_name='Server Mezz')
pivot_report.netport().to_excel(writer, sheet_name='Server Netport')
pivot_report.netcard().to_excel(writer, sheet_name='Server Netcard')
pivot_report.raid().to_excel(writer, sheet_name='Server RAID')
pivot_report.switchBoard().to_excel(writer, sheet_name='Switch Board')
pivot_report.storage().to_excel(writer, sheet_name='Storage')
pivot_report.storagedisk().to_excel(writer, sheet_name='Storage Disk')
pivot_report.network().to_excel(writer, sheet_name='Network Element')

writer.save()

current_time2 = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
print('End Time: {}'.format(current_time2))
stop = timeit.default_timer()
print('Total time to run the program: ', stop - start)
print("_________________________END_________________________")











'''
    Test_code
'''
# df_serverdetails = pd.read_json('asset_pivot_serverdetails.json')
# print(df_serverdetails)
# count_sites = len(df_serverdetails)
# print(count_sites)
# print(type(df_serverdetails))
#df_serverdetails.to_excel('asset_pivot_serverdetails.xlsx')
#for ind in df_serverdetails.index:
# print(len(df_serverdetails.columns))

# def flat_data(serverdetails):
#     count_blade = []
#     server_details = []
#     no_sites = len(serverdetails)
#     print(no_sites)
#     for i in range(no_sites):
#         no_blades = len(serverdetails[i])
#         count_blade.append(no_blades)
#         for j in range(no_blades):
#             out = serverdetails[i][j]
#             serverdetails.append(out)
#     print(count_blade)
#     serverdetails_json = json.dumps(serverdetails[12:], indent=4)
#     return flat_data
#
#
# json_data = open('asset_pivot_serverdetails.json', 'r')
# serverdetails = json.load(json_data)
# print(serverdetails)
# out = flat_data(serverdetails)
# df_ServerDetails = pd.read_json('out.json')
# print(df_ServerDetails)

#out = flat_data(serverdetails)
#df_ServerDetails = pd.read_json(out) #converts json to dataframe
#print(df_ServerDetails)
#
# FIELDS = ["location", "name", "ipAddress", "dn"]
# #ServerDetails = df_ServerDetails(jsondata)
# df_cpu = pd.json_normalize(df_ServerDetails[0], "CPU", FIELDS, record_prefix='_')
# df_cpu.to_excel('cpu.xlsx')
#
# data = open('asset_pivot_serverdetails.json', 'r')
# df_data = pd.read_json(data)
# print(df_data)





